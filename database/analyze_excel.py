import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\Codes\travis\smartsheet\Storage Materials.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")
print()

# Color mapping based on user description:
# Yellow - for quotation
# Green - already quoted  
# Red - needs clarification
# Brown - ongoing project

COLOR_STATUS = {
    'FFFFFF00': 'Yellow - Quotation',
    'FFFF00': 'Yellow - Quotation',
    'FF00FF00': 'Green - Already Quoted',
    '00FF00': 'Green - Already Quoted',
    'FFFF0000': 'Red - Needs Clarification',
    'FF0000': 'Red - Needs Clarification',
    '00EA352E': 'Red - Needs Clarification',
    'EA352E': 'Red - Needs Clarification',
    '00C6E7C8': 'Green - Already Quoted',
    'C6E7C8': 'Green - Already Quoted',
    '00D190DA': 'Purple/Pink',
    'D190DA': 'Purple/Pink',
    '00974C00': 'Brown - Ongoing Project',
    '974C00': 'Brown - Ongoing Project',
    '00FEFF85': 'Yellow - Quotation',
    'FEFF85': 'Yellow - Quotation',
}

# Find all project rows - projects are Level 0 outline rows that have names
# OR rows with colors that indicate project status
projects = []
all_colors = {}

for row in range(2, ws.max_row + 1):
    cell_a = ws.cell(row, 1)
    project_name = cell_a.value
    
    if not project_name or not isinstance(project_name, str):
        continue
    
    # Skip known non-project rows
    skip_names = ['Suppliers', 'Contractors', 'FINAL DRAWINGS', 'Door Order', 
                  'Construction Schedule', 'Bay Insulation', 'Central States',
                  'Whirlwind', 'Nucor', 'Canglong', 'Eric Escamilla', 'Tara Moggs',
                  'Eliseo Ruiz', 'Jose Porras', 'CR Menn', 'Cr Menn', 'Schulte Building',
                  'Balcon Builders', 'Certificate Of Insurance', 'Marketing Material',
                  'Contents for Marketing', 'Blank Storage', 'vendor/customer',
                  'w9 form', 'PEMB Form', 'Important Information', 'MBS/MBS Mini',
                  'Contracts', 'Wire Instructions', 'Competitor Quote', 'Hard Steel',
                  'National Steel', 'Tilt wall', 'Installer', 'Metal Increase',
                  'MPI Job', 'Sales Tax', 'Smartsheet User']
    
    is_skip = False
    for skip in skip_names:
        if skip.lower() in project_name.lower():
            is_skip = True
            break
    
    if is_skip:
        continue
    
    # Get outline level
    outline_level = ws.row_dimensions[row].outline_level
    
    # Get fill color
    fill_color = None
    if cell_a.fill and cell_a.fill.fgColor:
        rgb = cell_a.fill.fgColor.rgb
        if rgb and rgb != "00000000":
            fill_color = rgb
            if rgb not in all_colors:
                all_colors[rgb] = []
            all_colors[rgb].append(project_name[:40])
    
    # Get other columns
    location = ws.cell(row, 2).value
    project_address = ws.cell(row, 3).value
    project_label = ws.cell(row, 4).value
    quote_sent = ws.cell(row, 5).value
    reached_out = ws.cell(row, 6).value
    total_cogs = ws.cell(row, 7).value
    customer = ws.cell(row, 8).value
    phone = ws.cell(row, 9).value
    email = ws.cell(row, 10).value
    build_size = ws.cell(row, 11).value
    zip_code = ws.cell(row, 13).value
    project_type = ws.cell(row, 14).value
    project_sqft = ws.cell(row, 15).value
    our_quote_material = ws.cell(row, 19).value
    sales_tax = ws.cell(row, 20).value
    our_quote_with_tax = ws.cell(row, 21).value
    quote_accepted = ws.cell(row, 23).value
    deposit_paid = ws.cell(row, 24).value
    job_status = ws.cell(row, 34).value
    comments = ws.cell(row, 35).value
    
    # A project is likely if:
    # 1. It's outline level 0 (parent row)
    # 2. OR it has a color
    # 3. OR it has customer/location data
    # 4. AND the name looks like a location (contains state abbreviation or city name)
    
    is_project = False
    
    # Check if name looks like a project (location-based naming)
    import re
    state_pattern = r'\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|NB|Canada)\b'
    if re.search(state_pattern, project_name):
        is_project = True
    
    # Or if it starts with # (numbered project)
    if project_name.startswith('#'):
        is_project = True
    
    # Or if it has customer data
    if customer:
        is_project = True
    
    # Or if outline level 0 and has some data
    if outline_level == 0 and (location or project_label or fill_color):
        is_project = True
    
    if is_project:
        color_status = COLOR_STATUS.get(fill_color, None) if fill_color else None
        projects.append({
            'row': row,
            'name': project_name,
            'location': location,
            'address': project_address,
            'label': project_label,
            'customer': customer,
            'phone': phone,
            'email': email,
            'color': fill_color,
            'color_status': color_status,
            'outline_level': outline_level,
            'quote_sent': bool(quote_sent),
            'reached_out': bool(reached_out),
            'total_cogs': total_cogs,
            'build_size': build_size,
            'zip_code': zip_code,
            'project_type': project_type,
            'project_sqft': project_sqft,
            'our_quote_material': our_quote_material,
            'sales_tax': sales_tax,
            'our_quote_with_tax': our_quote_with_tax,
            'quote_accepted': quote_accepted,
            'deposit_paid': deposit_paid,
            'job_status': job_status,
            'comments': comments,
        })

print(f"Total projects found: {len(projects)}")
print()

print("All colors found:")
for color, names in sorted(all_colors.items(), key=lambda x: -len(x[1])):
    status = COLOR_STATUS.get(color, 'Unknown')
    print(f"  {color} ({status}): {len(names)} items")
    for n in names[:2]:
        print(f"    - {n}")
print()

print("Projects by status/color:")
by_status = {}
for p in projects:
    status = p['color_status'] or p['label'] or 'No Status'
    if status not in by_status:
        by_status[status] = []
    by_status[status].append(p['name'])

for status, names in sorted(by_status.items(), key=lambda x: -len(x[1])):
    print(f"  {status}: {len(names)}")
print()

print("First 50 projects:")
for i, p in enumerate(projects[:50]):
    status = p['color_status'] or p['label'] or '-'
    print(f"  {i+1}. {p['name'][:45]} | {p['customer'] or '-'} | {status}")

# Save to JSON for use in the app
with open('projects_data.json', 'w') as f:
    json.dump(projects, f, indent=2, default=str)
