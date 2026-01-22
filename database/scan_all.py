import openpyxl
import json
import re

wb = openpyxl.load_workbook(r'd:\Codes\travis\smartsheet\app\database\Storage Materials.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
print()

# Column mapping (1-indexed)
COL_PROJECT_NAME = 1      # A
COL_LOCATION = 2          # B
COL_ADDRESS = 3           # C
COL_LABEL = 4             # D - Project label
COL_QUOTE_SENT = 5        # E
COL_REACHED_OUT = 6       # F
COL_TOTAL_COGS = 7        # G
COL_CUSTOMER = 8          # H
COL_PHONE = 9             # I
COL_EMAIL = 10            # J
COL_BUILD_SIZE = 11       # K
COL_RECEIVED_DETAILS = 12 # L
COL_ZIP = 13              # M
COL_PROJECT_TYPE = 14     # N
COL_SQFT = 15             # O
COL_RECEIVED_QUOTE = 16   # P
COL_ERECTING_LABOR = 17   # Q
COL_CONCRETE_LABOR = 18   # R
COL_QUOTE_MATERIAL = 19   # S
COL_SALES_TAX = 20        # T
COL_QUOTE_WITH_TAX = 21   # U
COL_DELIVERED_QUOTE = 22  # V
COL_QUOTE_ACCEPTED = 23   # W
COL_DEPOSIT = 24          # X
COL_DRAWINGS_STATUS = 25  # Y
COL_EST_METAL_DATE = 26   # Z
COL_DOOR_ORDER_DATE = 27  # AA
COL_EST_DOOR_DATE = 28    # AB
COL_METAL_PROD = 29       # AC
COL_METAL_DELIVERY = 30   # AD
COL_DOOR_DELIVERY = 31    # AE
COL_FINAL_ACH = 32        # AF
COL_CONTRACTOR_DATE = 33  # AG
COL_JOB_STATUS = 34       # AH
COL_COMMENTS = 35         # AI

# Skip patterns - these are NOT projects
SKIP_PATTERNS = [
    r'^suppliers?$',
    r'^contractors?$',
    r'^final drawings',
    r'^door order',
    r'^construction schedule',
    r'^bay insulation',
    r'^central states',
    r'^whirlwind',
    r'^nucor',
    r'^canglong',
    r'^eric escamilla',
    r'^tara moggs',
    r'^eliseo ruiz',
    r'^jose porras',
    r'^cr menn',
    r'^schulte building',
    r'^balcon builders',
    r'^certificate of insurance',
    r'^marketing material',
    r'^contents for marketing',
    r'^blank storage',
    r'^vendor/customer',
    r'^w9 form',
    r'^pemb form',
    r'^important information',
    r'^mbs/mbs mini',
    r'^contracts?$',
    r'^wire instructions',
    r'^competitor quote',
    r'^hard steel',
    r'^national steel',
    r'^tilt wall',
    r'^installer$',
    r'^metal increase',
    r'^mpi job',
    r'^sales tax',
    r'^smartsheet user',
    r'^accounting$',
    r'^plans$',
    r'^takeoffs',
    r'^estimating',
    r'^business_development',
    r'^prebid',
    r'^prime_contract',
    r'^design_engineering',
    r'^arteras',
    r'^\d{2}_',  # Numbered folders like 01_, 02_
    r'^pemb unit cost',
    r'new project template',
]

def is_skip_row(name):
    if not name:
        return True
    name_lower = name.lower().strip()
    for pattern in SKIP_PATTERNS:
        if re.search(pattern, name_lower):
            return True
    return False

def get_cell_color(cell):
    """Get the fill color of a cell"""
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        if rgb and rgb != "00000000":
            return rgb
    return None

# Color status mapping
def get_color_status(color):
    if not color:
        return None
    color_map = {
        '00EA352E': 'Red - Needs Clarification',
        '00C6E7C8': 'Green - Already Quoted',
        '00D190DA': 'Purple/Pink',
        '00974C00': 'Brown - Ongoing Project',
        '00FEFF85': 'Yellow - Quotation',
        '00FEFF00': 'Yellow - Quotation',
        '0040B14B': 'Green - Already Quoted',
        '00237F2E': 'Green - Already Quoted',
        '007ED085': 'Green - Already Quoted',
        '00D0AF8F': 'Brown - Ongoing Project',
        '00592C00': 'Brown - Ongoing Project',
        '00FF8D00': 'Orange',
        '00EA5000': 'Orange',
    }
    return color_map.get(color)

projects = []
seen_names = set()

print("Scanning all rows for projects...")
print()

for row in range(2, ws.max_row + 1):
    project_name = ws.cell(row, COL_PROJECT_NAME).value
    
    if not project_name or not isinstance(project_name, str):
        continue
    
    project_name = project_name.strip()
    
    if is_skip_row(project_name):
        continue
    
    # Get all column values
    location = ws.cell(row, COL_LOCATION).value
    address = ws.cell(row, COL_ADDRESS).value
    label = ws.cell(row, COL_LABEL).value
    quote_sent = ws.cell(row, COL_QUOTE_SENT).value
    reached_out = ws.cell(row, COL_REACHED_OUT).value
    total_cogs = ws.cell(row, COL_TOTAL_COGS).value
    customer = ws.cell(row, COL_CUSTOMER).value
    phone = ws.cell(row, COL_PHONE).value
    email = ws.cell(row, COL_EMAIL).value
    build_size = ws.cell(row, COL_BUILD_SIZE).value
    zip_code = ws.cell(row, COL_ZIP).value
    project_type = ws.cell(row, COL_PROJECT_TYPE).value
    sqft = ws.cell(row, COL_SQFT).value
    quote_material = ws.cell(row, COL_QUOTE_MATERIAL).value
    sales_tax = ws.cell(row, COL_SALES_TAX).value
    quote_with_tax = ws.cell(row, COL_QUOTE_WITH_TAX).value
    quote_accepted = ws.cell(row, COL_QUOTE_ACCEPTED).value
    deposit = ws.cell(row, COL_DEPOSIT).value
    drawings_status = ws.cell(row, COL_DRAWINGS_STATUS).value
    est_metal_date = ws.cell(row, COL_EST_METAL_DATE).value
    door_order_date = ws.cell(row, COL_DOOR_ORDER_DATE).value
    est_door_date = ws.cell(row, COL_EST_DOOR_DATE).value
    metal_prod = ws.cell(row, COL_METAL_PROD).value
    metal_delivery = ws.cell(row, COL_METAL_DELIVERY).value
    door_delivery = ws.cell(row, COL_DOOR_DELIVERY).value
    contractor_date = ws.cell(row, COL_CONTRACTOR_DATE).value
    job_status = ws.cell(row, COL_JOB_STATUS).value
    comments = ws.cell(row, COL_COMMENTS).value
    
    # Get color
    color = get_cell_color(ws.cell(row, COL_PROJECT_NAME))
    color_status = get_color_status(color)
    
    # Determine if this is a project row
    # A project has: customer OR location OR label OR color OR looks like a location name
    is_project = False
    
    # Has customer data
    if customer:
        is_project = True
    
    # Has location data
    if location:
        is_project = True
    
    # Has project label
    if label:
        is_project = True
    
    # Has meaningful color
    if color_status:
        is_project = True
    
    # Name contains state abbreviation (looks like a project location)
    state_pattern = r'\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|Canada|NB)\b'
    if re.search(state_pattern, project_name):
        is_project = True
    
    # Name starts with # (numbered project)
    if project_name.startswith('#'):
        is_project = True
    
    # Has financial data (quote, cogs, etc.)
    if total_cogs or quote_with_tax or quote_material:
        is_project = True
    
    if not is_project:
        continue
    
    # Avoid duplicates (same name)
    name_key = project_name.lower()
    if name_key in seen_names:
        # But allow if this row has more data
        pass
    seen_names.add(name_key)
    
    projects.append({
        'row': row,
        'name': project_name,
        'location': location,
        'address': address,
        'label': label,
        'quote_sent': bool(quote_sent),
        'reached_out': bool(reached_out),
        'total_cogs': total_cogs if isinstance(total_cogs, (int, float)) else None,
        'customer': customer,
        'phone': str(phone) if phone else None,
        'email': email,
        'build_size': str(build_size) if build_size else None,
        'zip_code': str(zip_code) if zip_code else None,
        'project_type': project_type,
        'project_sqft': sqft if isinstance(sqft, (int, float)) else None,
        'our_quote_material': quote_material if isinstance(quote_material, (int, float)) else None,
        'sales_tax': sales_tax if isinstance(sales_tax, (int, float)) else None,
        'our_quote_with_tax': quote_with_tax if isinstance(quote_with_tax, (int, float)) else None,
        'quote_accepted': str(quote_accepted) if quote_accepted else None,
        'deposit_paid': deposit,
        'drawings_status': drawings_status,
        'est_metal_date': str(est_metal_date) if est_metal_date else None,
        'door_order_date': str(door_order_date) if door_order_date else None,
        'est_door_date': str(est_door_date) if est_door_date else None,
        'metal_production': metal_prod,
        'metal_delivery': metal_delivery,
        'door_delivery': door_delivery,
        'contractor_date': str(contractor_date) if contractor_date else None,
        'job_status': job_status,
        'comments': str(comments)[:500] if comments else None,
        'color': color,
        'color_status': color_status,
    })

print(f"Total projects found: {len(projects)}")
print()

# Stats
with_customer = sum(1 for p in projects if p['customer'])
with_location = sum(1 for p in projects if p['location'])
with_label = sum(1 for p in projects if p['label'])
with_color = sum(1 for p in projects if p['color_status'])
with_quote = sum(1 for p in projects if p['our_quote_with_tax'])

print(f"Projects with customer: {with_customer}")
print(f"Projects with location: {with_location}")
print(f"Projects with label: {with_label}")
print(f"Projects with color status: {with_color}")
print(f"Projects with quote: {with_quote}")
print()

# By label
print("By Label:")
label_counts = {}
for p in projects:
    lbl = p['label'] or 'No Label'
    label_counts[lbl] = label_counts.get(lbl, 0) + 1
for lbl, cnt in sorted(label_counts.items(), key=lambda x: -x[1]):
    print(f"  {lbl}: {cnt}")
print()

# By color status
print("By Color Status:")
color_counts = {}
for p in projects:
    cs = p['color_status'] or 'No Color'
    color_counts[cs] = color_counts.get(cs, 0) + 1
for cs, cnt in sorted(color_counts.items(), key=lambda x: -x[1]):
    print(f"  {cs}: {cnt}")
print()

# Sample projects
print("First 30 projects:")
for i, p in enumerate(projects[:30]):
    status = p['color_status'] or p['label'] or '-'
    cust = p['customer'] or '-'
    print(f"  {i+1}. Row {p['row']}: {p['name'][:40]} | {cust[:20]} | {status}")

# Save to JSON
with open('projects_data.json', 'w', encoding='utf-8') as f:
    json.dump(projects, f, indent=2, default=str)

print()
print(f"Saved {len(projects)} projects to projects_data.json")
